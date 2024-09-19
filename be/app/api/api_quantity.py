from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy import text
from sqlalchemy.orm import Session
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from io import BytesIO
from collections import defaultdict

from app.db.base import get_db

router = APIRouter()

def query_sew(date: date, fac: str, db: Session = Depends(get_db)):
    try:
        query = f"""
            select (LTRIM(RTRIM(workline))) AS workline,Style_No,SUM(JG_Count) as Qty
            from ETSDB_NAM.dbo.t_jjb
            where CAST(EndDate AS DATE) = '{date}' and left(workline,1) = '{fac}'
            and Gx_No = 99 and code not in ('88888','99999')
            group by CAST(EndDate AS DATE),(LTRIM(RTRIM(workline))),Style_No
            order by workline;
        """

        result = db.execute(text(query)).fetchall()
        data = [dict(row._mapping) for row in result]
        return data
    except:
        HTTPException(status_code=500, detail="Internal server error")

def query_iron(date: date, fac: str, db: Session = Depends(get_db)):
    try:
        query = f"""
            select SUM(JG_Count) as Qty
            from ETSDB_NAM.dbo.t_jjb
            where CAST(EndDate AS DATE) = '{date}'
            and Gx_No = {106 if fac == '1' else 102} and code not in ('88888','99999')
            group by CAST(EndDate AS DATE)
        """

        result = db.execute(text(query)).scalar()
        return result
    except:
        HTTPException(status_code=500, detail="Internal server error")


def query_qc(date: date, fac: str, db: Session = Depends(get_db)):
    try:
        query = f"""
            select SUM(JG_Count) as Qty
            from ETSDB_NAM.dbo.t_jjb
            where CAST(EndDate AS DATE) = '{date}'
            and Gx_No = {107 if fac == '1' else 103} and code not in ('88888','99999')
            group by CAST(EndDate AS DATE)
        """

        result = db.execute(text(query)).scalar()
        return result
    except:
        HTTPException(status_code=500, detail="Internal server error")
    
def query_pack(date: date, fac: str, db: Session = Depends(get_db)):
    try:
        query = f"""
            select SUM(JG_Count) as Qty
            from ETSDB_NAM.dbo.t_jjb
            where CAST(EndDate AS DATE) = '{date}'
            and Gx_No = {108 if fac == '1' else 104} and code not in ('88888','99999')
            group by CAST(EndDate AS DATE)
        """

        result = db.execute(text(query)).scalar()
        return result
    except:
        HTTPException(status_code=500, detail="Internal server error")
    

def query_data(date: date, fac: str, db: Session = Depends(get_db)):
    try:
        data_sew = query_sew(date, fac, db)
        data_iron = query_iron(date, fac, db)
        data_qc = query_qc(date, fac, db)
        data_pack = query_pack(date, fac, db)
        return {
            "sew": data_sew,
            "iron": data_iron,
            "qc": data_qc,
            "pack": data_pack
        }
    except:
        HTTPException(status_code=500, detail="Internal server error")

@router.get("/")
def get_data(date: date, fac: str, db: Session = Depends(get_db)): 
    try:
        sew, iron, qc, pack = query_data(date, fac, db).values()
        return {"sew": sew, "iron": iron, "qc": qc, "pack": pack}
    except:
        HTTPException(status_code=500, detail="Internal server error")

@router.get("/excel")
def export_excel(date: date, fac: str, db: Session = Depends(get_db)):
    try:
        sew, iron, qc, pack = query_data(date, fac, db).values()

        quantity_totals = defaultdict(int)
        sum_quantity = 0
        for item in sew:
            quantity_totals[item['workline']] += item['Qty']
            sum_quantity += item['Qty']

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        headers = ["STT", "Line", "Style", "Output (hệ thống)", "Tổng sản lượng"]
        ws.append(headers)

        # Fill data
        index = 0
        current_row = 2
        for i, item in enumerate(sew):
            if item["workline"] != sew[i - 1]["workline"]: index += 1
            ws.cell(row=current_row, column=1, value=index)
            ws.cell(row=current_row, column=2, value=item['workline'])
            ws.cell(row=current_row, column=3, value=item['Style_No'])
            ws.cell(row=current_row, column=4, value=item['Qty'])
            ws.cell(row=current_row, column=5, value=quantity_totals[item['workline']])
            current_row += 1

        # Merge row
        current_merge_start = 2 
        for i in range(3, len(sew) + 3): 
            if i == len(sew) + 2 or ws.cell(i, 2).value != ws.cell(i - 1, 2).value:
                if current_merge_start < i - 1:
                    ws.merge_cells(start_row=current_merge_start, start_column=1, end_row=i - 1, end_column=1)
                    ws.merge_cells(start_row=current_merge_start, start_column=2, end_row=i - 1, end_column=2)
                    ws.merge_cells(start_row=current_merge_start, start_column=5, end_row=i - 1, end_column=5)
                current_merge_start = i

        # Add sum row
        bold_font = Font(bold=True)
        sumCol = ws.cell(row=current_row, column=1, value="SUM")
        ws.cell(row=current_row, column=4, value=sum_quantity)
        ws.cell(row=current_row, column=5, value=sum_quantity)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        sumCol.font = bold_font
        current_row += 1

        # Add O: Other direct labour
        O = [
            {"name": "Là 106", "quantity": iron}, 
            {"name": "QC Là 107", "quantity": qc}, 
            {"name": "Đóng gói 108", "quantity": pack}
        ]
        for item in O:
            index += 1
            ws.cell(row=current_row, column=1, value=index)
            ws.cell(row=current_row, column=2, value=item.get("name"))
            ws.cell(row=current_row, column=4, value=item.get("quantity", 0))
            ws.cell(row=current_row, column=5, value=item.get("quantity", 0))
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
            current_row += 1

        # Style
        # Define border style
        thin_border = Border(left=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        top=Side(border_style="thin"),
                        bottom=Side(border_style="thin"))
        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in ws.iter_rows(min_row=1, max_row=len(sew) + 5, max_col=5):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return Response(
            buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=sanluong_{}.xlsx".format(date)}
        )
    except:
        HTTPException(status_code=500, detail="Internal server error")

@router.get("/excel_secretary")
def export_excel_for_secretary(date: date, fac: str, db: Session = Depends(get_db)):
    try:
        sew, iron, qc, pack = query_data(date, fac, db).values()

        quantity_totals = defaultdict(int)
        sum_quantity = 0
        for item in sew:
            quantity_totals[item['workline']] += item['Qty']
            sum_quantity += item['Qty']

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        headers = ["STT", "Line", "Style", "Output (hệ thống)", "Tổng sản lượng", "Sản lượng thực tế", "+/-"]
        ws.append(headers)

        # Fill data
        index = 0
        current_row = 2
        for i, item in enumerate(sew):
            if item["workline"] != sew[i - 1]["workline"]: index += 1
            ws.cell(row=current_row, column=1, value=index)
            ws.cell(row=current_row, column=2, value=item['workline'])
            ws.cell(row=current_row, column=3, value=item['Style_No'])
            ws.cell(row=current_row, column=4, value=item['Qty'])
            ws.cell(row=current_row, column=5, value=quantity_totals[item['workline']])
            current_row += 1

        # Merge row
        current_merge_start = 2 
        for i in range(3, len(sew) + 3): 
            if i == len(sew) + 2 or ws.cell(i, 2).value != ws.cell(i - 1, 2).value:
                if current_merge_start < i - 1:
                    ws.merge_cells(start_row=current_merge_start, start_column=1, end_row=i - 1, end_column=1)
                    ws.merge_cells(start_row=current_merge_start, start_column=2, end_row=i - 1, end_column=2)
                    ws.merge_cells(start_row=current_merge_start, start_column=5, end_row=i - 1, end_column=5)
                    ws.merge_cells(start_row=current_merge_start, start_column=6, end_row=i - 1, end_column=6)
                    ws.merge_cells(start_row=current_merge_start, start_column=7, end_row=i - 1, end_column=7)
                    ws.cell(row=current_merge_start, column=7).value = f"=IF(ISBLANK(F{current_merge_start}),-E{current_merge_start},E{current_merge_start}-F{current_merge_start})"
                else:
                    row = i - 1
                    ws.cell(row=row, column=7).value = f"=IF(ISBLANK(F{row}),-E{row},E{row}-F{row})"
                current_merge_start = i


        # Add sum row
        bold_font = Font(bold=True)
        ws.cell(row=current_row, column=1, value="TOT")
        ws.cell(row=current_row, column=4, value=sum_quantity)
        ws.cell(row=current_row, column=5, value=sum_quantity)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        ws.cell(row=current_row, column=6).value = f"=SUM(F2:F{current_row - 1})"
        ws.cell(row=current_row, column=7).value = f"=IF(ISBLANK(F{current_row}),-E{current_row},E{current_row}-F{current_row})"
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).font = bold_font
        current_row += 1

        # Add O: Other direct labour
        O = [
            {"name": "Là 106", "quantity": iron}, 
            {"name": "QC Là 107", "quantity": qc}, 
            {"name": "Đóng gói 108", "quantity": pack}
        ]
        for item in O:
            index += 1
            ws.cell(row=current_row, column=1, value=index)
            ws.cell(row=current_row, column=2, value=item.get("name"))
            ws.cell(row=current_row, column=4, value=item.get("quantity", 0))
            ws.cell(row=current_row, column=5, value=item.get("quantity", 0))
            ws.cell(row=current_row, column=6, value=item.get("quantity", 0))
            ws.cell(row=current_row, column=7, value=0)
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
            current_row += 1

        # Style
        # Define border style
        thin_border = Border(left=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        top=Side(border_style="thin"),
                        bottom=Side(border_style="thin"))
        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in ws.iter_rows(min_row=1, max_row=len(sew) + 5, max_col=7):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        ws.column_dimensions['G'].width = 10 

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return Response(
            buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=sanluong(2)_{}.xlsx".format(date)}
        )
    except: 
        HTTPException(status_code=500, detail="Internal server error")